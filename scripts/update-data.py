"""Update food price data from MAFF public files.

Run from the project root:
  python scripts/update-data.py

The script downloads the latest public MAFF vegetable workbook and
meat/egg PDF, parses the national average prices, and replaces the
`const foods = [...]` block at the top of script.js.
"""

from __future__ import annotations

import json
import re
import sys
import tempfile
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl
import pypdf


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_JS = ROOT / "script.js"

VEG_PAGE = "https://www.maff.go.jp/j/zyukyu/anpo/kouri/k_yasai/h22index.html"
MEAT_PAGE = "https://www.maff.go.jp/j/zyukyu/anpo/kouri/k_gyuniku/index.html"

VEG_META = {
    "キャベツ": ("cabbage", "#2f7d57", "#e9f5ee"),
    "ねぎ": ("negi", "#6d9b4f", "#eff6e8"),
    "レタス": ("lettuce", "#4f9d69", "#edf7ef"),
    "ばれいしょ": ("potato", "#9a7a3a", "#f5efe3"),
    "たまねぎ": ("onion", "#b48120", "#fbf3dd"),
    "きゅうり": ("cucumber", "#2f8e72", "#e7f5f1"),
    "トマト": ("tomato", "#bf4f43", "#faece8"),
    "にんじん": ("carrot", "#c76b35", "#fbefe8"),
    "はくさい": ("hakusai", "#8aa23f", "#f3f6e6"),
    "だいこん": ("daikon", "#6b8a5b", "#eef4eb"),
}

MEAT_META = [
    ("輸入牛肉", "beef-import", "#8c493f", "#f5e9e6", "円/100g"),
    ("国産牛肉", "beef-domestic", "#7c3f36", "#f3e7e5", "円/100g"),
    ("豚肉ロース", "pork", "#3d6f95", "#e8f1f7", "円/100g"),
    ("鶏もも肉", "chicken", "#7c5b2f", "#f4eee4", "円/100g"),
    ("鶏卵", "egg", "#6b6f3f", "#f1f3df", "円/10個"),
]


def fetch_text(url: str) -> str:
    with urllib.request.urlopen(url, timeout=30) as response:
        return response.read().decode("utf-8", errors="replace")


def download(url: str, target: Path) -> None:
    with urllib.request.urlopen(url, timeout=60) as response:
        target.write_bytes(response.read())


def latest_href(page_url: str, pattern: str) -> str:
    html = fetch_text(page_url)
    matches = re.findall(r'href="([^"]+%s[^"]*)"' % pattern, html)
    if not matches:
        raise RuntimeError(f"No link matching {pattern!r} found on {page_url}")
    return urllib.parse.urljoin(page_url, matches[0])


def numeric(value):
    return isinstance(value, (int, float))


def comment_for(name: str, parity: float, previous_ratio: float | None) -> str:
    if parity <= 0.96:
        base = "平年より安めで、買い時寄りです。"
    elif parity >= 1.12:
        base = "高値圏のため、必要分だけの購入が無難です。"
    elif parity >= 1.06:
        base = "平年よりやや高めです。"
    else:
        base = "ほぼ平年並みです。"

    if previous_ratio is None:
        trend = ""
    elif previous_ratio >= 1.03:
        trend = "直近は値上がりしています。"
    elif previous_ratio <= 0.97:
        trend = "直近は値下がりしています。"
    else:
        trend = "直近は横ばいです。"
    return f"公式最新値では平年比{round(parity * 100)}%です。{base}{trend}"


def parse_vegetables(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    price_ws = workbook["価格"]
    week_ws = workbook["前週比"]
    parity_ws = workbook["平年比"]
    headers = [cell.value for cell in price_ws[2]]

    rows = list(price_ws.iter_rows(values_only=True))
    latest_index = max(
        i for i, row in enumerate(rows)
        if isinstance(row[0], str) and row[0].startswith("令和") and any(numeric(v) for v in row[1:11])
    )
    latest_row = rows[latest_index]
    source_date = latest_row[0]
    week_row = list(week_ws.iter_rows(values_only=True))[latest_index]
    parity_row = list(parity_ws.iter_rows(values_only=True))[latest_index]

    foods = []
    for col in range(1, 11):
        name = headers[col]
        price = latest_row[col]
        parity = parity_row[col]
        week_ratio = week_row[col]
        if name not in VEG_META or not numeric(price) or not numeric(parity):
            continue

        history = []
        for row in rows[max(2, latest_index - 6): latest_index + 1]:
            if numeric(row[col]):
                history.append(round(row[col]))
        if len(history) < 2:
            history = [round(price / week_ratio), round(price)] if numeric(week_ratio) else [round(price), round(price)]

        item_id, accent, tint = VEG_META[name]
        foods.append({
            "id": item_id,
            "name": name,
            "category": "vegetable",
            "unit": "円/kg",
            "accent": accent,
            "tint": tint,
            "reference": round(price / parity),
            "sourceDate": source_date,
            "comment": comment_for(name, parity, week_ratio if numeric(week_ratio) else None),
            "history": history,
        })
    return foods


def parse_meat(path: Path) -> list[dict]:
    reader = pypdf.PdfReader(path)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    source_match = re.search(r"令和[0-9０-９]+年[0-9０-９]+月\n?（[^）]+）", text)
    source_date = source_match.group(0).replace("\n", "") if source_match else "最新公表月"

    table = re.search(r"価格\s+([0-9 ]+)\n前月比\s+([0-9% ]+)\n平年比\s+([0-9% ]+)", text)
    if not table:
        raise RuntimeError("Could not parse meat/egg summary table from PDF.")
    prices = [int(v) for v in table.group(1).split()]
    month_ratios = [int(v.rstrip("%")) / 100 for v in table.group(2).split()]
    parity_ratios = [int(v.rstrip("%")) / 100 for v in table.group(3).split()]

    numbers = []
    page2 = reader.pages[1].extract_text() or ""
    for line in page2.splitlines():
        vals = [int(v) for v in re.findall(r"\b[0-9]{3}\b", line)]
        if len(vals) == 5:
            numbers.append(vals)
    histories_by_col = list(zip(*numbers[:8])) if numbers else [[p] for p in prices]

    foods = []
    for index, (name, item_id, accent, tint, unit) in enumerate(MEAT_META):
        price = prices[index]
        parity = parity_ratios[index]
        month_ratio = month_ratios[index]
        history = list(reversed([int(v) for v in histories_by_col[index]]))
        if history[-1] != price:
            history.append(price)
        foods.append({
            "id": item_id,
            "name": name,
            "category": "meat",
            "unit": unit,
            "accent": accent,
            "tint": tint,
            "reference": round(price / parity),
            "sourceDate": source_date,
            "comment": f"公式最新値では平年比{round(parity * 100)}%、前月比{round(month_ratio * 100)}%です。",
            "history": history[-8:],
        })
    return foods


def js_value(value) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def replace_foods_block(foods: list[dict]) -> None:
    source = SCRIPT_JS.read_text(encoding="utf-8")
    replacement = "const foods = " + js_value(foods) + ";"
    updated = re.sub(r"const foods = \[[\s\S]*?\];", replacement, source, count=1)
    if updated == source:
        raise RuntimeError("Could not find foods block in script.js")
    SCRIPT_JS.write_text(updated, encoding="utf-8")


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        vegetable_url = latest_href(VEG_PAGE, r"\.xlsx")
        meat_url = latest_href(MEAT_PAGE, r"\.pdf")
        vegetable_path = tmpdir / "vegetables.xlsx"
        meat_path = tmpdir / "meat.pdf"
        print(f"Downloading vegetables: {vegetable_url}")
        download(vegetable_url, vegetable_path)
        print(f"Downloading meat/egg: {meat_url}")
        download(meat_url, meat_path)

        foods = parse_vegetables(vegetable_path) + parse_meat(meat_path)
        replace_foods_block(foods)
        print(f"Updated {SCRIPT_JS.name} with {len(foods)} items.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
